#!/usr/bin/env python3
"""
Combine multiple LDAK heritability TSV files into a single Excel table.

This script reads all ${output_prefix}.*.ldak.tsv files and combines them into
a formatted Excel table with hierarchical column headers:
- Sheet 1 (Heritability): Summary table with all phenotypes
  - Phenotype (first column)
  - Heritability | split | SNP/INDEL/SV/SNP_INDEL/SNP_INDEL_SV
  - Heritability | Unsplit | SNP/INDEL/SV/SNP_INDEL/SNP_INDEL_SV
- Sheet 2 (SNP_INDEL): Detailed REML information for SNP_INDEL merged type
  - Both split and unsplit data side by side
  - Converged status
  - Her_K1 (Heritability, SE)
  - Her_K2 (Heritability, SE)
  - Her_All (Heritability, SE)
- Sheet 3 (SNP_INDEL_SV): Detailed REML information for SNP_INDEL_SV merged type
  - Both split and unsplit data side by side
  - Converged status
  - Her_K1 (Heritability, SE)
  - Her_K2 (Heritability, SE)
  - Her_K3 (Heritability, SE)
  - Her_All (Heritability, SE)
"""

import sys
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def parse_way(filename, output_prefix):
    """Extract way type from filename like 'prefix.SNP_split.ldak.tsv'"""
    basename = os.path.basename(filename)
    # Remove output_prefix and .ldak.tsv
    way_part = basename.replace(f"{output_prefix}.", "").replace(".ldak.tsv", "")
    return way_part

def read_tsv_file(filepath):
    """Read a single TSV file and return DataFrame"""
    try:
        df = pd.read_csv(filepath, sep='\t')
        return df
    except Exception as e:
        print(f"Warning: Could not read {filepath}: {e}", file=sys.stderr)
        return None

def parse_reml_file(filepath):
    """Parse LDAK REML file and extract key information including all columns"""
    try:
        data = {}
        with open(filepath, 'r') as f:
            lines = f.readlines()
        
        # Parse header information
        for line in lines:
            line = line.strip()
            if line.startswith('Converged'):
                parts = line.split()
                data['Converged'] = parts[1] if len(parts) > 1 else 'NA'
            elif line.startswith('Component'):
                # This is the header line for component data
                continue
            elif line.startswith('Her_K1'):
                parts = line.split()
                if len(parts) >= 6:
                    data['Her_K1_Heritability'] = float(parts[1]) if parts[1] != 'NA' else None
                    data['Her_K1_SE'] = float(parts[2]) if parts[2] != 'NA' else None
                    data['Her_K1_Size'] = float(parts[3]) if parts[3] != 'NA' else None
                    data['Her_K1_Mega_Intensity'] = float(parts[4]) if parts[4] != 'NA' else None
                    data['Her_K1_Mega_Intensity_SE'] = float(parts[5]) if parts[5] != 'NA' else None
            elif line.startswith('Her_K2'):
                parts = line.split()
                if len(parts) >= 6:
                    data['Her_K2_Heritability'] = float(parts[1]) if parts[1] != 'NA' else None
                    data['Her_K2_SE'] = float(parts[2]) if parts[2] != 'NA' else None
                    data['Her_K2_Size'] = float(parts[3]) if parts[3] != 'NA' else None
                    data['Her_K2_Mega_Intensity'] = float(parts[4]) if parts[4] != 'NA' else None
                    data['Her_K2_Mega_Intensity_SE'] = float(parts[5]) if parts[5] != 'NA' else None
            elif line.startswith('Her_K3'):
                parts = line.split()
                if len(parts) >= 6:
                    data['Her_K3_Heritability'] = float(parts[1]) if parts[1] != 'NA' else None
                    data['Her_K3_SE'] = float(parts[2]) if parts[2] != 'NA' else None
                    data['Her_K3_Size'] = float(parts[3]) if parts[3] != 'NA' else None
                    data['Her_K3_Mega_Intensity'] = float(parts[4]) if parts[4] != 'NA' else None
                    data['Her_K3_Mega_Intensity_SE'] = float(parts[5]) if parts[5] != 'NA' else None
            elif line.startswith('Her_Top'):
                parts = line.split()
                if len(parts) >= 6:
                    data['Her_Top_Heritability'] = float(parts[1]) if parts[1] != 'NA' else None
                    data['Her_Top_SE'] = float(parts[2]) if parts[2] != 'NA' else None
                    data['Her_Top_Size'] = float(parts[3]) if parts[3] != 'NA' else None
                    data['Her_Top_Mega_Intensity'] = float(parts[4]) if parts[4] != 'NA' else None
                    data['Her_Top_Mega_Intensity_SE'] = float(parts[5]) if parts[5] != 'NA' else None
            elif line.startswith('Her_All'):
                parts = line.split()
                if len(parts) >= 6:
                    data['Her_All_Heritability'] = float(parts[1]) if parts[1] != 'NA' else None
                    data['Her_All_SE'] = float(parts[2]) if parts[2] != 'NA' else None
                    data['Her_All_Size'] = float(parts[3]) if parts[3] != 'NA' else None
                    data['Her_All_Mega_Intensity'] = float(parts[4]) if parts[4] != 'NA' else None
                    data['Her_All_Mega_Intensity_SE'] = float(parts[5]) if parts[5] != 'NA' else None
        
        return data
    except Exception as e:
        print(f"Warning: Could not parse REML file {filepath}: {e}", file=sys.stderr)
        return None

def find_reml_files(directory, way_type):
    """Find all REML files for a specific way type (e.g., SNP_INDEL_split or SNP_INDEL_SV_unsplit)"""
    reml_files = {}
    
    # Pattern: directory/way_type/results/phenotype_name/phenotype_name_LDAK-Thin.reml
    pattern = os.path.join(directory, way_type, "results", "*", "*_LDAK-Thin.reml")
    files = glob.glob(pattern)
    
    for filepath in files:
        # Extract phenotype from the directory name (parent directory of the REML file)
        # Path structure: .../way_type/results/phenotype_name/phenotype_name_LDAK-Thin.reml
        phenotype_dir = os.path.basename(os.path.dirname(filepath))
        reml_files[phenotype_dir] = filepath
    
    return reml_files

def add_reml_detail_sheet(wb, directory, way_type, sheet_name, phenotypes):
    """Add a sheet with detailed REML information for a specific way type"""
    
    # Find all REML files for this way type
    reml_files = find_reml_files(directory, way_type)
    
    if not reml_files:
        print(f"Warning: No REML files found for {way_type}", file=sys.stderr)
        return
    
    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)
    
    # Define headers with all columns from REML file
    headers = ['Phenotype', 'Converged', 
               'Her_K1', 'Her_K1_SE', 'Her_K1_Size', 'Her_K1_Mega_Intensity', 'Her_K1_Mega_Intensity_SE',
               'Her_K2', 'Her_K2_SE', 'Her_K2_Size', 'Her_K2_Mega_Intensity', 'Her_K2_Mega_Intensity_SE',
               'Her_K3', 'Her_K3_SE', 'Her_K3_Size', 'Her_K3_Mega_Intensity', 'Her_K3_Mega_Intensity_SE',
               'Her_Top', 'Her_Top_SE', 'Her_Top_Size', 'Her_Top_Mega_Intensity', 'Her_Top_Mega_Intensity_SE',
               'Her_All', 'Her_All_SE', 'Her_All_Size', 'Her_All_Mega_Intensity', 'Her_All_Mega_Intensity_SE']
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data for each phenotype
    row_idx = 2
    for phenotype in phenotypes:
        if phenotype in reml_files:
            reml_data = parse_reml_file(reml_files[phenotype])
            if reml_data:
                ws.cell(row=row_idx, column=1, value=phenotype)
                ws.cell(row=row_idx, column=2, value=reml_data.get('Converged', 'NA'))
                
                # Her_K1 (5 columns)
                ws.cell(row=row_idx, column=3, value=reml_data.get('Her_K1_Heritability'))
                ws.cell(row=row_idx, column=4, value=reml_data.get('Her_K1_SE'))
                ws.cell(row=row_idx, column=5, value=reml_data.get('Her_K1_Size'))
                ws.cell(row=row_idx, column=6, value=reml_data.get('Her_K1_Mega_Intensity'))
                ws.cell(row=row_idx, column=7, value=reml_data.get('Her_K1_Mega_Intensity_SE'))
                
                # Her_K2 (5 columns)
                ws.cell(row=row_idx, column=8, value=reml_data.get('Her_K2_Heritability'))
                ws.cell(row=row_idx, column=9, value=reml_data.get('Her_K2_SE'))
                ws.cell(row=row_idx, column=10, value=reml_data.get('Her_K2_Size'))
                ws.cell(row=row_idx, column=11, value=reml_data.get('Her_K2_Mega_Intensity'))
                ws.cell(row=row_idx, column=12, value=reml_data.get('Her_K2_Mega_Intensity_SE'))
                
                # Her_K3 (5 columns)
                ws.cell(row=row_idx, column=13, value=reml_data.get('Her_K3_Heritability'))
                ws.cell(row=row_idx, column=14, value=reml_data.get('Her_K3_SE'))
                ws.cell(row=row_idx, column=15, value=reml_data.get('Her_K3_Size'))
                ws.cell(row=row_idx, column=16, value=reml_data.get('Her_K3_Mega_Intensity'))
                ws.cell(row=row_idx, column=17, value=reml_data.get('Her_K3_Mega_Intensity_SE'))
                
                # Her_Top (5 columns)
                ws.cell(row=row_idx, column=18, value=reml_data.get('Her_Top_Heritability'))
                ws.cell(row=row_idx, column=19, value=reml_data.get('Her_Top_SE'))
                ws.cell(row=row_idx, column=20, value=reml_data.get('Her_Top_Size'))
                ws.cell(row=row_idx, column=21, value=reml_data.get('Her_Top_Mega_Intensity'))
                ws.cell(row=row_idx, column=22, value=reml_data.get('Her_Top_Mega_Intensity_SE'))
                
                # Her_All (5 columns)
                ws.cell(row=row_idx, column=23, value=reml_data.get('Her_All_Heritability'))
                ws.cell(row=row_idx, column=24, value=reml_data.get('Her_All_SE'))
                ws.cell(row=row_idx, column=25, value=reml_data.get('Her_All_Size'))
                ws.cell(row=row_idx, column=26, value=reml_data.get('Her_All_Mega_Intensity'))
                ws.cell(row=row_idx, column=27, value=reml_data.get('Her_All_Mega_Intensity_SE'))
                
                # Format numeric cells
                for col in range(3, 28):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000000'
                
                row_idx += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    
    # Set width for all data columns (C through AA for 27 columns total)
    for col_idx in range(3, 28):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
        ws.column_dimensions[col_letter].width = 14
    
    print(f"Added sheet '{sheet_name}' with {row_idx - 2} phenotypes", file=sys.stderr)

def add_reml_detail_sheet_combined(wb, directory, way_types, sheet_name, phenotypes):
    """Add a sheet with detailed REML information combining multiple way types (e.g., split and unsplit)
    
    Args:
        wb: Workbook object
        directory: Directory containing results
        way_types: List of way types to combine (e.g., ['SNP_INDEL_split', 'SNP_INDEL_unsplit'])
        sheet_name: Name for the sheet
        phenotypes: List of phenotypes
    """
    
    # Find all REML files for all way types
    all_reml_files = {}
    for way_type in way_types:
        reml_files = find_reml_files(directory, way_type)
        if reml_files:
            all_reml_files[way_type] = reml_files
    
    if not all_reml_files:
        print(f"Warning: No REML files found for any of {way_types}", file=sys.stderr)
        return
    
    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)
    
    # Dynamically detect which components are present in the REML files
    # Check a sample REML file to see which Her_K* components have data
    available_components = set()
    for way_type in all_reml_files:
        for phenotype, reml_file in all_reml_files[way_type].items():
            reml_data = parse_reml_file(reml_file)
            if reml_data:
                # Check which Her_K* components have non-None heritability values
                for comp in ['Her_K1', 'Her_K2', 'Her_K3']:
                    if reml_data.get(f'{comp}_Heritability') is not None:
                        available_components.add(comp)
                # Always include Her_All
                if reml_data.get('Her_All_Heritability') is not None:
                    available_components.add('Her_All')
            break  # Only check one file per way_type
        if available_components:
            break  # Found components, no need to check more
    
    # Build components list in order: Her_K1, Her_K2, Her_K3, Her_All
    components = []
    for comp in ['Her_K1', 'Her_K2', 'Her_K3']:
        if comp in available_components:
            components.append(comp)
    if 'Her_All' in available_components:
        components.append('Her_All')
    
    if not components:
        print(f"Warning: No valid components found in REML files for {sheet_name}", file=sys.stderr)
        return
    
    print(f"Sheet '{sheet_name}' will include components: {components}", file=sys.stderr)
    
    # Define headers with hierarchical structure
    # Row 1: Main headers
    ws.cell(row=1, column=1, value='Phenotype')
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    
    col_offset = 2
    for way_type in way_types:
        # Extract suffix (split or unsplit)
        suffix = 'split' if 'split' in way_type.lower() and 'unsplit' not in way_type.lower() else 'Unsplit'
        
        # Calculate number of columns: Converged (1) + components * 2 (Heritability + SE)
        num_cols = 1 + len(components) * 2
        
        # Merge cells for way type header
        ws.cell(row=1, column=col_offset, value=suffix)
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset + num_cols - 1)
        
        # Row 2: Sub-headers for each way type
        sub_headers = ['Converged']
        for comp in components:
            sub_headers.append(comp)
            sub_headers.append(f'{comp}_SE')
        
        for idx, header in enumerate(sub_headers):
            cell = ws.cell(row=2, column=col_offset + idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_offset += num_cols
    
    # Format main headers
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, col_offset):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Format Phenotype header
    phenotype_cell = ws.cell(row=1, column=1)
    phenotype_cell.fill = header_fill
    phenotype_cell.font = header_font
    phenotype_cell.alignment = center_alignment
    
    # Write data for each phenotype
    row_idx = 3
    for phenotype in phenotypes:
        ws.cell(row=row_idx, column=1, value=phenotype)
        
        col_offset = 2
        num_cols = 1 + len(components) * 2
        
        # Convert phenotype name from "XXX.tsv" format to "processed_XXX" format for REML lookup
        # Remove .tsv suffix and add processed_ prefix
        phenotype_key = phenotype.replace('.tsv', '')
        phenotype_key = f'processed_{phenotype_key}'
        
        for way_type in way_types:
            if way_type in all_reml_files and phenotype_key in all_reml_files[way_type]:
                reml_data = parse_reml_file(all_reml_files[way_type][phenotype_key])
                if reml_data:
                    # Converged
                    ws.cell(row=row_idx, column=col_offset, value=reml_data.get('Converged', 'NA'))
                    
                    # Write data for each component
                    current_col = col_offset + 1
                    for comp in components:
                        # Heritability
                        ws.cell(row=row_idx, column=current_col, value=reml_data.get(f'{comp}_Heritability'))
                        # SE
                        ws.cell(row=row_idx, column=current_col + 1, value=reml_data.get(f'{comp}_SE'))
                        current_col += 2
                    
                    # Format numeric cells
                    for col in range(col_offset + 1, col_offset + num_cols):
                        cell = ws.cell(row=row_idx, column=col)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.000000'
            else:
                # Fill with empty cells if data not available
                for col in range(col_offset, col_offset + num_cols):
                    ws.cell(row=row_idx, column=col, value='')
            
            col_offset += num_cols
        
        row_idx += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    
    # Helper function to convert column index to Excel column letter
    def col_idx_to_letter(col_idx):
        """Convert column index (1-based) to Excel column letter"""
        letter = ''
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    for col_idx in range(2, col_offset):
        col_letter = col_idx_to_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14
    
    print(f"Added combined sheet '{sheet_name}' with {row_idx - 3} phenotypes and {len(way_types)} way types", file=sys.stderr)

def combine_heritability_files(directory, output_prefix, output_file):
    """Combine all heritability TSV files into Excel table
    
    Args:
        directory: Directory containing results structure (with way/results/phenotype subdirs)
        output_prefix: Prefix for output files
        output_file: Output Excel file path
    """
    
    # Find all matching TSV files in current directory (staged by Nextflow)
    # Try current directory first (Nextflow stages files here)
    pattern = f"{output_prefix}.*.ldak.tsv"
    tsv_files = glob.glob(pattern)
    
    # If not found in current directory, try the provided directory
    if not tsv_files:
        pattern = os.path.join(directory, f"{output_prefix}.*.ldak.tsv")
        tsv_files = glob.glob(pattern)
    
    # If still not found, try heritability directory
    if not tsv_files:
        pattern = os.path.join("heritability", f"{output_prefix}.*.ldak.tsv")
        tsv_files = glob.glob(pattern)
    
    if not tsv_files:
        print(f"Error: No TSV files found matching pattern {output_prefix}.*.ldak.tsv", file=sys.stderr)
        print(f"Searched in current directory, {directory}, and heritability/", file=sys.stderr)
        sys.exit(1)
    
    print(f"Found {len(tsv_files)} TSV files to combine", file=sys.stderr)
    
    # Read all TSV files and store by way type
    data_dict = {}  # way -> DataFrame
    all_phenotypes = set()
    
    for tsv_file in tsv_files:
        way = parse_way(tsv_file, output_prefix)
        df = read_tsv_file(tsv_file)
        if df is not None and not df.empty:
            data_dict[way] = df
            all_phenotypes.update(df['Phenotype'].tolist())
    
    if not data_dict:
        print("Error: No valid data found in any TSV file", file=sys.stderr)
        sys.exit(1)
    
    # Define the order of way types (matching the image)
    way_types = ['SNP', 'INDEL', 'SV', 'SNP_INDEL', 'SNP_INDEL_SV']
    split_types = [f"{wt}_split" for wt in way_types]
    unsplit_types = [f"{wt}_unsplit" for wt in way_types]
    
    # Create ordered list of phenotypes
    sorted_phenotypes = sorted(all_phenotypes)
    
    # Build data structure for Excel output
    # We'll create a structure that matches the hierarchical header format
    data_rows = []
    
    # Process each phenotype
    for phenotype in sorted_phenotypes:
        row = [phenotype]  # Start with Phenotype column
        
        # Process split types: only Heritability (no P_value)
        for way in split_types:
            if way in data_dict:
                df = data_dict[way]
                pheno_data = df[df['Phenotype'] == phenotype]
                if not pheno_data.empty:
                    row.append(pheno_data.iloc[0]['Heritability'])
                else:
                    row.append(None)
            else:
                row.append(None)
        
        # Process unsplit types: only Heritability (no P_value)
        for way in unsplit_types:
            if way in data_dict:
                df = data_dict[way]
                pheno_data = df[df['Phenotype'] == phenotype]
                if not pheno_data.empty:
                    row.append(pheno_data.iloc[0]['Heritability'])
                else:
                    row.append(None)
            else:
                row.append(None)
        
        data_rows.append(row)
    
    # Calculate mean row
    mean_row = ['mean']
    for col_idx in range(1, len(data_rows[0])):  # Skip Phenotype column
        col_values = [row[col_idx] for row in data_rows if row[col_idx] is not None and pd.notna(row[col_idx])]
        if col_values:
            mean_val = sum(col_values) / len(col_values)
            mean_row.append(mean_val)
        else:
            mean_row.append(None)
    
    data_rows.append(mean_row)
    
    # Create Excel file with hierarchical headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Heritability"
    
    # Write headers (2 rows) - no P_value columns
    # Row 1: Main header
    ws.cell(row=1, column=1, value='Phenotype')
    
    # Calculate positions for split section (only Heritability, no P_value)
    split_start_col = 2
    split_end_col = split_start_col + len(split_types) - 1
    
    ws.cell(row=1, column=split_start_col, value='Heritability')
    ws.merge_cells(start_row=1, start_column=split_start_col, end_row=1, end_column=split_end_col)
    
    unsplit_start_col = split_end_col + 1
    unsplit_end_col = unsplit_start_col + len(unsplit_types) - 1
    
    ws.cell(row=1, column=unsplit_start_col, value='Heritability')
    ws.merge_cells(start_row=1, start_column=unsplit_start_col, end_row=1, end_column=unsplit_end_col)
    
    # Row 2: split/Unsplit sub-headers
    ws.cell(row=2, column=split_start_col, value='split')
    ws.merge_cells(start_row=2, start_column=split_start_col, end_row=2, end_column=split_end_col)
    
    ws.cell(row=2, column=unsplit_start_col, value='Unsplit')
    ws.merge_cells(start_row=2, start_column=unsplit_start_col, end_row=2, end_column=unsplit_end_col)
    
    # Row 3: Individual way types (one column per type, no P_value)
    col = split_start_col
    for way_type in way_types:
        ws.cell(row=3, column=col, value=way_type)
        col += 1
    
    col = unsplit_start_col
    for way_type in way_types:
        ws.cell(row=3, column=col, value=way_type)
        col += 1
    
    # Write data rows (starting from row 4)
    for row_idx, data_row in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Format numeric cells - all are Heritability values now
            if isinstance(value, (int, float)) and pd.notna(value) and col_idx > 1:
                cell.number_format = '0.000000'
    
    # Format headers
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
    
    # Merge Phenotype header across first 3 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Add detailed REML sheets for merged types
    # Sheet 2: SNP_INDEL (both split and unsplit)
    # Sheet 3: SNP_INDEL_SV (both split and unsplit)
    add_reml_detail_sheet_combined(wb, directory, 
                                   ['SNP_INDEL_split', 'SNP_INDEL_unsplit'], 
                                   'SNP_INDEL', sorted_phenotypes)
    
    add_reml_detail_sheet_combined(wb, directory, 
                                   ['SNP_INDEL_SV_split', 'SNP_INDEL_SV_unsplit'], 
                                   'SNP_INDEL_SV', sorted_phenotypes)
    
    # Save file
    wb.save(output_file)
    
    print(f"Successfully created Excel file: {output_file}", file=sys.stderr)

def main():
    if len(sys.argv) != 4:
        print("Usage: combine_heritability_table.py <directory> <output_prefix> <output_file>", file=sys.stderr)
        sys.exit(1)
    
    directory = sys.argv[1]
    output_prefix = sys.argv[2]
    output_file = sys.argv[3]
    
    if not os.path.isdir(directory):
        print(f"Error: Directory {directory} does not exist", file=sys.stderr)
        sys.exit(1)
    
    combine_heritability_files(directory, output_prefix, output_file)

if __name__ == '__main__':
    main()

